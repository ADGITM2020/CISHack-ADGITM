from pathlib import Path

import openpyxl

from bill.models import Product
from saraswati_enterprises import settings

serial = 0
file_path = Path(settings.BASE_DIR, 'bill/scripts/products.xlsx')


def product_id(product_name):
    global serial
    product_name = f"SARA{product_name[:3]}{serial:04d}"
    serial += 1
    return product_name


def update_serial_number():
    global serial
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb["Sheet1"]
    serial = ws.max_row

def update_product_list():
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb["Sheet1"]
    max_row = ws.max_row



def run(*args):
    assert args[0]
    global serial
    serial = int(args[0])
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb["Sheet1"]
    for i in range(1, ws.max_row + 1):
        try:
            ws.cell(row=i, column=7).value = product_id(ws[i][0].value)
            product, created = Product.objects.get_or_create(product_id=ws[i][6].value,
                                                             product_name=ws[i][0].value.strip(),
                                                             hsn_code=ws[i][1].value,
                                                             product_price=ws[i][2].value,
                                                             product_cgst=ws[i][3].value, product_sgst=ws[i][4].value,
                                                             product_igst=ws[i][5].value)
            print(created)
        except Exception as e:
            print(e)
    wb.save(file_path)

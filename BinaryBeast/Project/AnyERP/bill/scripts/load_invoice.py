import re
from datetime import datetime
from pathlib import Path

import openpyxl
from django.db.models import Q

from bill.models import Invoice, Company, Vehicle, Product, InvoiceProduct
from bill.scripts.load_products import product_id, update_serial_number
from saraswati_enterprises import settings


def check_date(invoice_date, po_date):
    if type(po_date) == str and type(invoice_date) == str:
        po_date = datetime.strptime(po_date, "%d.%m.%Y")
        invoice_date = datetime.strptime(invoice_date, "%d.%m.%Y")
    elif type(po_date) == str:
        po_date = datetime.strptime(po_date, "%d.%m.%Y")
        invoice_date = invoice_date.date()
    elif type(invoice_date) == str:
        invoice_date = datetime.strptime(invoice_date, "%d.%m.%Y")
        po_date = po_date.date()
    else:
        po_date = po_date.date()
        invoice_date = invoice_date.date()
    return invoice_date, po_date


def check_eway(eway_uncleaned):
    try:
        eway = eway_uncleaned.split(".")[1].replace(" ", "")
    except:
        eway = "N/A"
    return eway


def create_product(name, hsn, price, gst, igst=False):
    product_obj = Product.objects.get_or_create(
        product_price=price,
        product_name=name,
        hsn_code=hsn,
        product_id=product_id(name),
    )
    if igst:
        product_obj.product_igst = gst

    product_obj.product_sgst = gst
    product_obj.product_cgst = gst
    product_obj.save()
    print(product_obj)
    return product_obj


def get_product_object(product_title, product_price):
    product_obj = Product.objects.filter(
        Q(product_name=product_title) & Q(product_price=product_price) | Q(product_name=product_title))
    if product_obj.exists():
        return product_obj[0]
    print("returning false")
    return False


def convert_excel_to_sql_data(file_path):
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    ws = wb["Sheet1"]
    invoice_text = ws[7][1].value
    regex_pattern = r" ?/ ?(?P<invoice_no>[\d]+)"
    regex = re.compile(regex_pattern)
    match = regex.search(invoice_text)
    try:
        Invoice.objects.get(invoice_no=match["invoice_no"])
        return None
    except Invoice.DoesNotExist:
        print("Creating Invoice", match['invoice_no'])
    gst = ws[12][2].value.split(":")[-1].strip()
    invoice_term = ws[12][7].value
    cartage = ws[30][7].value
    if ws[10][6].value.upper() == "BY HAND":
        driver_object = Vehicle.objects.get(vehicle_no=ws[10][6].value.upper())
    else:
        driver_object = Vehicle.objects.get(vehicle_no=ws[10][6].value.replace(" ", ""))
    invoice_company_object = Company.objects.get(gstin=gst)

    sum_total = 0
    invoice_igst = ws[32][6].value
    invoice_cgst = ws[33][6].value
    invoice_sgst = ws[34][6].value
    invoice_date, po_date = check_date(ws[7][7].value, ws[9][7].value)
    ref_transport_name = ws[36][1].value
    ref_transport_no = ws[35][3].value
    invoice = Invoice(invoice_no=match["invoice_no"],
                      bill_to=invoice_company_object,
                      payment_term=invoice_term,
                      po_no=ws[9][6].value,
                      po_date=po_date,
                      bill_date=invoice_date,
                      cartage=ws[30][7].value,
                      invoice_driver=driver_object,
                      eway=check_eway(ws[28][2].value),
                      invoice_igst=invoice_igst,
                      invoice_cgst=invoice_cgst,
                      invoice_sgst=invoice_sgst,
                      year=invoice_date.year,
                      )
    if ref_transport_name and ref_transport_no:
        invoice.ref_transport_no = ref_transport_no
        invoice.ref_transport_name = ref_transport_name
    elif ref_transport_name:
        invoice.ref_transport_name = ref_transport_name
    invoice.save()
    for i in range(14, ws.max_row):
        if ws[i][2].value and ws[i][3].value:
            product_name = ws[i][2].value.strip()
            print(product_name)
            if "POLY VINYL" in product_name:
                product_name = "POLY VINYL ALCOHAL " + product_name.split(" ", 3)[3].replace(" ", "")
            product = get_product_object(product_name, ws[i][6].value)
            if ws[i + 1][2].value and ws[i + 1][3].value:
                print("hello")
                if product is False:
                    print("creating new product")
                    if invoice_igst:
                        product = create_product(product_name, ws[i][3].value, ws[i][6].value, invoice_igst, igst=True)
                    else:
                        product = create_product(product_name, ws[i][3].value, ws[i][6].value, invoice_cgst)
                invoice_product = InvoiceProduct(
                    bill_no=invoice,
                    item=product,
                    quantity=ws[i][5].value,
                    packing=ws[i][4].value,
                    rate=ws[i][6].value,
                    total=ws[i][7].value,
                )
                sum_total += ws[i][7].value
                invoice_product.save()
            else:
                invoice_product = InvoiceProduct(
                    bill_no=invoice,
                    item=product,
                    quantity=ws[i][5].value,
                    packing=ws[i][4].value,
                    rate=ws[i][6].value,
                    total=ws[i][7].value,
                    description=ws[i + 1][2].value
                )
                sum_total += ws[i][7].value
                invoice_product.save()
        else:
            break
    grand_total = sum_total + cartage
    invoice.sum_total = sum_total
    invoice.grand_total = grand_total
    if invoice_cgst:
        invoice.grand_total_final = round((grand_total * 2 * invoice_cgst) + grand_total)
    else:
        invoice.grand_total_final = round((grand_total * 2 * invoice_igst) + grand_total)
    invoice.save()
    wb.close()
    return "successfull"


def run(*args):
    max_bills = int(args[0])
    update_serial_number()
    for i in range(3, max_bills + 1):
        print(i)
        file_path = Path(settings.BASE_DIR, f'bill/old_bills/INV2019-20/INV{i:04d}.xlsm')
        if file_path.exists():
            status = convert_excel_to_sql_data(file_path)
            if status:
                print(status)
            else:
                print("already exit")
        else:
            print("file doesn't exists")

import os

import openpyxl

from bill.models import Company
from saraswati_enterprises import settings


def run():
    file_path = os.path.join(settings.BASE_DIR, 'bill/scripts/company.xlsx')
    wb = openpyxl.load_workbook(file_path)
    ws = wb["Sheet1"]
    for i in range(2, ws.max_row + 1):
        state_code = ws[i][4].value
        try:
            company, created = Company.objects.get_or_create(company_name=ws[i][0].value.lstrip(),
                                                             address=ws[i][1].value.lstrip(),
                                                             location=ws[i][2].value.lstrip(),
                                                             gstin=ws[i][3].value.split(" : ")[1].lstrip(),
                                                             state_code=f"{state_code:02d}".lstrip())
            if not created:
                print("failed for",ws[i][0].value.lstrip())
            else:
                print("success")
        except:
            print(ws[i][3].value)

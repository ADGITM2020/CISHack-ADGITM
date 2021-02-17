import os

import openpyxl

from bill.models import Vehicle
from saraswati_enterprises import settings

#  --script-args for passing aruguments
def run():
    file_path = os.path.join(settings.BASE_DIR, 'bill/scripts/transport.xlsx')
    wb = openpyxl.load_workbook(file_path)
    ws = wb["Sheet1"]
    for i in range(1, ws.max_row + 1):
        try:
            vehicle, created = Vehicle.objects.get_or_create(vehicle_driver=ws[i][0].value, vehicle_no=ws[i][1].value)
            print(created)
        except Exception as e:
            print(e)

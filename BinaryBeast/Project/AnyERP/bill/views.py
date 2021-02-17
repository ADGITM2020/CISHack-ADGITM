import os
from decimal import Decimal

import convertapi
import openpyxl
from django.conf import settings
from django.contrib.auth import logout
from django.contrib.auth.decorators import login_required
from django.forms import formset_factory
from django.http import HttpResponse
from django.shortcuts import render, redirect
from django.utils.decorators import method_decorator
from django.views.generic import ListView
from openpyxl.styles import Alignment

from .drive_manager import get_link
from .forms import InvoiceForm, ProductForm
from .models import Vehicle, InvoiceProduct, Invoice
from .num2word import convert_to_words


def invoice_detail(request):
    return render(request, "bill/detail.html", {})


def logout_view(request):
    logout(request)
    return redirect('home')


# vehicle registration form


def vehicle_register(request):
    if request.method == 'GET':
        return render(request, 'bill/vehicle_form.html')


# create view of vehicle without the get request. There is no template needed as the call is asynchronous


def create_vehicle(request):
    if request.method == 'POST':
        vehicle_driver = request.POST['vehicle_driver']
        vehicle_no = request.POST['vehicle_no']
        Vehicle.objects.create(
            vehicle_driver=vehicle_driver, vehicle_no=vehicle_no)
        print("created successfully")
        return HttpResponse('')
    return render(request, 'bill/vehicle_form.html')


@login_required
def invoice_view(request):
    context = {}
    ProductFormSet = formset_factory(ProductForm)
    form = InvoiceForm()

    data = {
        'form-INITIAL_FORMS': 0,
        'form-TOTAL_FORMS': 1,
        'form-MIN_NUM_FORMS': 0,
        'form-MAX_NUM_FORMS': 10,
    }
    formset = ProductFormSet(data, prefix="form")
    context['formset'] = formset
    context['form'] = form
    if request.method == 'POST':
        formset = ProductFormSet(request.POST)
        form = InvoiceForm(request.POST)
        if form.is_valid() and formset.is_valid():
            instance = form.save(commit=False)
            instance.save()
            s_no, sum_tot = 1, 0
            for product in formset:
                ans = product.save(commit=False)
                ans.bill_no = instance
                if ans.rate:
                    ans.total = ans.rate * ans.quantity
                else:
                    ans.rate = ans.item.product_price
                    ans.total = ans.rate * ans.quantity

                if s_no == 1 and instance.bill_to.state_code != "06":
                    instance.invoice_igst = ans.item.product_igst
                else:
                    instance.invoice_cgst = ans.item.product_cgst
                    instance.invoice_sgst = ans.item.product_sgst
                s_no += 1
                sum_tot += ans.total
                ans.save()

            instance.sum_total = sum_tot
            instance.grand_total = sum_tot + instance.cartage
            igst_val = instance.grand_total * instance.invoice_igst * Decimal('0.01')
            cgst_val = instance.grand_total * instance.invoice_cgst * Decimal('0.01')
            sgst_val = cgst_val
            instance.year = instance.bill_date.year
            instance.grand_total_final = round(
                instance.grand_total + sgst_val + igst_val + cgst_val)
            instance.save()
            return redirect("bill:info", pk=instance.pk, year=instance.year)
    return render(request, "bill/invoice.html", context)


def bill(request, pk):
    try:
        invoice = Invoice.objects.get(pk=pk)
    except Invoice.DoesNotExist:
        return render(request, '404.html', status=404)
    if request.method == "POST":
        eway = request.POST['eway']
        invoice.eway = eway
        invoice.save()
        return redirect('bill:info', pk=pk)

    data = {}
    bill_products = InvoiceProduct.objects.filter(
        bill_no=invoice)
    data["bill_products"] = bill_products
    data["invoice"] = invoice
    return render(request, 'bill/info.html', data)


def download_excel(request, pk):
    invoice = Invoice.objects.get(pk=pk)
    output_file = os.path.join(settings.BASE_DIR, 'static/file/output.xlsx')
    with open(output_file, 'rb') as fh:
        response = HttpResponse(fh, content_type="application/vnd.ms-excel")
        response['Content-Disposition'] = f'attachment; filename=INV{invoice.year}-{invoice.invoice_no:04d}.xlsx'
        return response


def convert_pdf(request, pk):
    invoice = Invoice.objects.get(pk=pk)
    convertapi.api_secret = 'lR4Dkryb5oIGVUJr'
    file = os.path.join(settings.BASE_DIR, 'static/file/output.xlsx')
    output_file = os.path.join(settings.BASE_DIR, 'static/file')
    convertapi.convert('pdf', {
        'File': file,
    }, from_format='xlsx').save_files(output_file)
    pdf_path = os.path.join(settings.BASE_DIR, 'static/file/output.pdf')
    link = get_link(invoice.invoice_no, invoice.year, pdf_path)
    invoice.invoice_url = link
    invoice.save()
    return HttpResponse(link)


def download_pdf(request, pk):
    invoice = Invoice.objects.get(pk=pk)
    url = invoice.invoice_url
    if url:
        id = url.split("/d/")[-1].split("/", 1)[0]
        download_url = f"https://drive.google.com/u/6/uc?id={id}&export=download"
    return redirect(download_url)


def chart_view(request):
    return render(request, 'bill/chart.html', {})


def generate_files(request, pk):
    invoice = Invoice.objects.get(pk=pk)
    file_path = os.path.join(settings.BASE_DIR, 'static/file/example.xlsx')
    wb = openpyxl.load_workbook(file_path)
    ws = wb["Sheet1"]
    bill_products = InvoiceProduct.objects.filter(
        bill_no=invoice)
    s_no = 1
    for product in bill_products:
        equi_row = 14 + s_no
        ws.cell(row=equi_row, column=2, value=s_no)
        ws.cell(row=equi_row, column=3, value=product.item.product_name)
        ws.cell(row=equi_row, column=4, value=product.item.hsn_code)
        ws.cell(row=equi_row, column=5, value=product.packing)
        ws.cell(row=equi_row, column=6, value=product.quantity)
        ws.cell(row=equi_row, column=7, value=product.rate)
        ws.cell(row=equi_row, column=8, value=product.total)
        s_no += 1
    igst_val = invoice.grand_total * invoice.invoice_igst
    cgst_val = sgst_val = invoice.grand_total * invoice.invoice_cgst
    before_round = invoice.grand_total + sgst_val + igst_val + cgst_val
    ws["H28"].value = invoice.cartage
    ws["H34"].value = invoice.grand_total_final
    ws["G30"].value = invoice.invoice_igst
    ws["G31"].value = invoice.invoice_cgst
    ws["G32"].value = invoice.invoice_sgst
    ws["H33"].value = invoice.grand_total_final - before_round
    ws.merge_cells('C23:D24')
    ws['C23'] = invoice.grand_total_final
    ws['C23'].alignment = Alignment(horizontal='center')
    ws["H8"].value = invoice.bill_date
    ws["H10"].value = invoice.po_date
    if invoice.bill_date.month > 3:
        ws[
            "B8"].value = f"INVOICE NO. {invoice.bill_date.year}-{str(invoice.bill_date.year + 1)[2:]} / {invoice.invoice_no:03d}"
    else:
        ws[
            "B8"].value = f"INVOICE NO. {invoice.bill_date.year - 1}-{str(invoice.bill_date.year)[2:]} / {invoice.invoice_no:03d}"

    ws["G10"].value = invoice.po_no
    ws["G11"].value = invoice.invoice_driver.vehicle_no
    ws["H13"].value = invoice.payment_term

    ws["C10"].value = invoice.bill_to.company_name
    ws["C11"].value = invoice.bill_to.address
    ws["C12"].value = invoice.bill_to.location
    ws["C13"].value = invoice.bill_to.gstin
    ws["E13"].value = invoice.bill_to.state_code

    num_word = convert_to_words(int(invoice.grand_total_final))
    ws["B31"].value = num_word
    ws["B34"].value = invoice.ref_transport_name
    ws["D33"].value = invoice.ref_transport_no
    if invoice.eway:
        ws["C28"].value = 'EWAY NO. ' + invoice.eway
    else:
        ws["C28"].value = 'EWAY NO. N/A'

    output_file = os.path.join(settings.BASE_DIR, 'static/file/output.xlsx')
    wb.save(output_file)
    return HttpResponse("success")


@method_decorator(login_required, name="dispatch")
class BillSection(ListView):
    model = Invoice
    template_name = 'bill/bill-section.html'
    queryset = Invoice.objects.all()
    context_object_name = 'bills'
    ordering = '-bill_date'
    paginate_by = 10